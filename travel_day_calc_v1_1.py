import sys
import pandas as pd
import re
import json
import os
import logging
from datetime import timedelta
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QLabel,
    QPushButton, QCheckBox, QSpinBox, QComboBox,
    QTableWidget, QTableWidgetItem, QHeaderView, QMessageBox, QTextEdit, QFileDialog,
    QDialog, QDialogButtonBox, QDateEdit, QLineEdit, QListWidget, QListWidgetItem,
    QMenuBar, QMenu, QAction, QInputDialog
)
from PyQt5.QtCore import Qt, QDate
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Configure logging
logging.basicConfig(
    filename='travel_day_calc.log',
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# Constants
DEFAULT_CONFIG_FILE = 'default_config.json'

# Data Processing Functions
def load_and_process_schedule(file_path):
    """
    Loads the cast schedule from an Excel file, processes it by splitting cast members,
    cleaning their names, and sorting the data.

    Args:
        file_path (str): Path to the Excel file containing the cast schedule.

    Returns:
        pd.DataFrame: Processed DataFrame with 'Cast Member' and 'Shooting Date'.
    """
    try:
        df = pd.read_excel(file_path)
        logging.info(f"Successfully loaded Excel file: {file_path}")
    except Exception as e:
        logging.critical(f"Error reading Excel file: {e}")
        QMessageBox.critical(None, "File Read Error", f"Error reading Excel file:\n{e}")
        raise e

    # Split cast members into a list
    df['Cast List'] = df['CAST'].apply(split_cast)

    # Explode the cast list to have one cast member per row
    df_exploded = df.explode('Cast List')

    # Rename columns for clarity
    df_exploded.rename(columns={'Cast List': 'Cast Member', 'SHOOTING DATE': 'Shooting Date'}, inplace=True)

    # Parse 'Shooting Date' as datetime
    try:
        df_exploded['Shooting Date'] = pd.to_datetime(df_exploded['Shooting Date'], format='%d.%m.%Y')
        logging.info("Successfully parsed 'Shooting Date' as datetime.")
    except Exception as e:
        logging.critical(f"Error parsing shooting dates: {e}")
        QMessageBox.critical(None, "Date Parsing Error", f"Error parsing shooting dates:\n{e}")
        raise e

    # Drop rows without a cast member
    df_exploded.dropna(subset=['Cast Member'], inplace=True)

    # Clean cast member names
    df_exploded['Cast Member'] = df_exploded['Cast Member'].apply(clean_cast_member)

    # Remove duplicates
    df_cast_dates = df_exploded[['Cast Member', 'Shooting Date']].drop_duplicates()

    # Sort the DataFrame
    df_cast_dates.sort_values(['Cast Member', 'Shooting Date'], inplace=True)
    logging.info("Successfully processed cast dates.")

    return df_cast_dates

def split_cast(cast_str):
    """
    Splits the 'CAST' string into a list of individual cast members.

    Args:
        cast_str (str): Comma-separated string of cast members.

    Returns:
        list: List of individual cast members.
    """
    if pd.isnull(cast_str):
        return []
    else:
        cast_list = cast_str.split(',')
        cast_list = [cast_member.strip() for cast_member in cast_list]
        logging.debug(f"Split cast string '{cast_str}' into {cast_list}")
        return cast_list

def clean_cast_member(name):
    """
    Cleans individual cast member names by removing leading numbers and trailing information.

    Args:
        name (str): Raw cast member name.

    Returns:
        str: Cleaned cast member name.
    """
    if isinstance(name, str):
        name = re.sub(r'^\d+\.', '', name)  # Remove leading numbers and dots
        name = re.sub(r'\s*\(\d+\)', '', name)  # Remove trailing numbers in parentheses
        name = name.strip()
        logging.debug(f"Cleaned cast member name: '{name}'")
        return name
    else:
        logging.warning(f"Encountered non-string cast member name: {name}")
        return ''

# Configuration Management
def load_default_config():
    """
    Loads the default configuration from 'default_config.json'. If the file doesn't exist, initializes it with default values.

    Returns:
        dict: Default configuration dictionary containing 'home_locations'.
    """
    if os.path.exists(DEFAULT_CONFIG_FILE):
        with open(DEFAULT_CONFIG_FILE, 'r') as f:
            try:
                config = json.load(f)
                logging.info(f"Successfully loaded default configuration from {DEFAULT_CONFIG_FILE}")
            except json.JSONDecodeError as e:
                logging.critical(f"Error parsing '{DEFAULT_CONFIG_FILE}': {e}")
                QMessageBox.critical(None, "Configuration Error", f"Error parsing '{DEFAULT_CONFIG_FILE}':\n{e}")
                sys.exit(1)
    else:
        config = {
            "home_locations": ["Local", "Away"]  # Default locations
        }
        save_default_config(config)
        logging.info(f"Initialized new default configuration and saved to {DEFAULT_CONFIG_FILE}")
    return config

def save_default_config(config):
    """
    Saves the default configuration to 'default_config.json'.

    Args:
        config (dict): Configuration dictionary to save.
    """
    try:
        with open(DEFAULT_CONFIG_FILE, 'w') as f:
            json.dump(config, f, indent=4)
        logging.info(f"Default configuration saved to {DEFAULT_CONFIG_FILE}")
    except Exception as e:
        logging.critical(f"Error saving default configuration: {e}")
        QMessageBox.critical(None, "Save Error", f"Error saving default configuration:\n{e}")

def load_project_config(project_file):
    """
    Loads the project configuration from the specified file.

    Args:
        project_file (str): Path to the project file.

    Returns:
        dict: Project configuration dictionary.
    """
    with open(project_file, 'r') as f:
        project_config = json.load(f)

    # Convert date strings back to pd.Timestamp in shooting periods
    shooting_periods = project_config.get('config', {}).get('shooting_periods', [])
    for period in shooting_periods:
        if 'Start Date' in period and isinstance(period['Start Date'], str):
            period['Start Date'] = pd.to_datetime(period['Start Date'])
        if 'End Date' in period and isinstance(period['End Date'], str):
            period['End Date'] = pd.to_datetime(period['End Date'])

    logging.info(f"Successfully loaded project configuration from {project_file}")
    return project_config

def save_project_config(project_file, project_data):
    """
    Saves the project configuration to the specified file.

    Args:
        project_file (str): Path to the project file.
        project_data (dict): Project data to save.
    """
    try:
        # Define a custom JSON encoder
        class CustomEncoder(json.JSONEncoder):
            def default(self, obj):
                if isinstance(obj, pd.Timestamp):
                    return obj.strftime('%Y-%m-%dT%H:%M:%S')
                return json.JSONEncoder.default(self, obj)

        with open(project_file, 'w') as f:
            json.dump(project_data, f, indent=4, cls=CustomEncoder)
        logging.info(f"Project configuration saved to {project_file}")
    except Exception as e:
        logging.critical(f"Error saving project configuration: {e}")
        QMessageBox.critical(None, "Save Error", f"Error saving project configuration:\n{e}")

# Logic Functions
def calculate_gaps(df_cast_dates):
    """
    Calculates gaps between shooting dates for each cast member.

    Args:
        df_cast_dates (pd.DataFrame): DataFrame with 'Cast Member' and 'Shooting Date'.

    Returns:
        pd.DataFrame: DataFrame with additional columns for gaps and weekends in gaps.
    """
    df_cast_dates = df_cast_dates.copy()
    df_cast_dates['Previous Shooting Date'] = df_cast_dates.groupby('Cast Member')['Shooting Date'].shift(1)
    # Adjust gap calculation to exclude shooting days
    df_cast_dates['Gap'] = (df_cast_dates['Shooting Date'] - df_cast_dates['Previous Shooting Date']).dt.days - 1
    df_cast_dates['Gap'] = df_cast_dates['Gap'].fillna(0).astype(int)  # Handle NaN values
    df_cast_dates['Weekends in Gap'] = df_cast_dates.apply(weekends_in_gap, axis=1)
    logging.info("Calculated gaps and weekends in gaps.")
    return df_cast_dates

def weekends_in_gap(row):
    """
    Calculates the number of weekend days within a gap period.

    Args:
        row (pd.Series): Row containing 'Previous Shooting Date' and 'Shooting Date'.

    Returns:
        int: Number of weekend days in the gap.
    """
    if pd.isnull(row['Previous Shooting Date']) or row['Gap'] <= 0:
        return 0
    else:
        gap_range = pd.date_range(
            start=row['Previous Shooting Date'] + timedelta(days=1),
            end=row['Shooting Date'] - timedelta(days=1)
        )
        weekend_days = sum(gap_range.weekday >= 5)  # Saturday=5, Sunday=6
        logging.debug(f"Calculated {weekend_days} weekend days in gap for {row['Cast Member']}")
        return weekend_days

def apply_user_logic(df_cast_dates, max_gap, weekend_policy, arrival_option, departure_option,
                     cast_member_settings, shooting_periods):
    """
    Applies the user-defined logic to determine travel and accommodation needs for cast members.

    Args:
        df_cast_dates (pd.DataFrame): DataFrame with cast members and their shooting dates.
        max_gap (int): Maximum allowable gap between shooting days.
        weekend_policy (bool): Whether to send cast home on weekends.
        arrival_option (str): Arrival option ("Day Before Shooting" or "Same Day as Shooting").
        departure_option (str): Departure option ("Same Day as Shooting" or "Day After Shooting").
        cast_member_settings (dict): Settings for each cast member.
        shooting_periods (list): List of shooting periods with name, location, start date, and end date.

    Returns:
        tuple: (DataFrame summarizing results, list of export data dictionaries, list of calendar data)
    """
    df_filtered = df_cast_dates[df_cast_dates['Cast Member'].isin(cast_member_settings.keys())].copy()
    df_filtered = df_filtered.sort_values(['Cast Member', 'Shooting Date']).reset_index(drop=True)

    # Assign 'Period Name' and 'Shooting Location' to each shooting date
    def get_period_info(date):
        for period in shooting_periods:
            # Ensure 'Start Date' and 'End Date' are Timestamps
            if isinstance(period['Start Date'], str):
                try:
                    period['Start Date'] = pd.to_datetime(period['Start Date'])
                    if pd.isnull(period['Start Date']):
                        continue
                except Exception as e:
                    logging.critical(f"Error parsing Start Date for period '{period.get('Name', 'Unnamed')}': {e}")
                    QMessageBox.critical(None, "Date Parsing Error", f"Error parsing Start Date for period '{period.get('Name', 'Unnamed')}':\n{e}")
                    continue
            if isinstance(period['End Date'], str):
                try:
                    period['End Date'] = pd.to_datetime(period['End Date'])
                    if pd.isnull(period['End Date']):
                        continue
                except Exception as e:
                    logging.critical(f"Error parsing End Date for period '{period.get('Name', 'Unnamed')}': {e}")
                    QMessageBox.critical(None, "Date Parsing Error", f"Error parsing End Date for period '{period.get('Name', 'Unnamed')}':\n{e}")
                    continue
            # Debug: log the types
            logging.debug(f"Period '{period['Name']}' Start Date Type: {type(period['Start Date'])}, End Date Type: {type(period['End Date'])}")
            if period['Start Date'] <= date <= period['End Date']:
                return period['Name'], period['Location']
        return None, None

    # Apply the function and create two new columns
    df_filtered[['Period Name', 'Shooting Location']] = df_filtered['Shooting Date'].apply(
        lambda date: pd.Series(get_period_info(date))
    )
    df_filtered['Home Location'] = df_filtered['Cast Member'].map(
        lambda name: cast_member_settings[name]['home_location']
    )
    df_filtered['Include'] = df_filtered['Cast Member'].map(
        lambda name: cast_member_settings[name]['include']
    )

    # Keep only included cast members
    df_filtered = df_filtered[df_filtered['Include']]

    # Determine if cast member requires travel based on home location and shooting location
    def requires_travel(row):
        if pd.isnull(row['Shooting Location']):
            return False  # No specific shooting location; assume no travel
        elif row['Home Location'].lower() == row['Shooting Location'].lower():
            return False  # Same location; no travel
        else:
            return True  # Different location; travel required

    df_filtered['Requires Travel'] = df_filtered.apply(requires_travel, axis=1)

    summary = []
    results = []
    export_data = []
    calendar_data = []

    # Group by 'Cast Member', 'Home Location', 'Period Name', and 'Requires Travel'
    grouping_columns = ['Cast Member', 'Home Location', 'Period Name', 'Requires Travel']
    for group_keys, group in df_filtered.groupby(grouping_columns):
        name, home_location, period_name, requires_travel = group_keys
        group = group.sort_values('Shooting Date').reset_index(drop=True)

        # Calculate gaps between shooting dates
        group['Gap'] = group['Shooting Date'].diff().dt.days - 1
        group['Gap'] = group['Gap'].fillna(0).astype(int)

        # Identify new periods based on max_gap
        group['New Period'] = (group['Gap'] >= max_gap).cumsum()

        periods = group.groupby('New Period')

        total_accommodation = 0
        total_travel = 0

        for period_number, period_group in periods:
            period_length = len(period_group)

            if requires_travel:
                # Initialize travel dates set
                travel_dates = set()

                # Determine arrival and departure dates
                if arrival_option == 'Day Before Shooting':
                    arrival_date = period_group['Shooting Date'].min() - timedelta(days=1)
                    travel_dates.add(arrival_date)
                elif arrival_option == 'Same Day as Shooting':
                    arrival_date = period_group['Shooting Date'].min()
                    travel_dates.add(arrival_date)

                if departure_option == 'Day After Shooting':
                    departure_date = period_group['Shooting Date'].max() + timedelta(days=1)
                    travel_dates.add(departure_date)
                elif departure_option == 'Same Day as Shooting':
                    departure_date = period_group['Shooting Date'].max()
                    # Only add departure date if it's different from arrival date
                    if departure_date != arrival_date:
                        travel_dates.add(departure_date)

                # Number of unique travel dates
                period_travel = len(travel_dates)

                # Calculate accommodation nights
                period_accommodation = (departure_date - arrival_date).days

                # Calculate gap days within the period
                shooting_dates = period_group['Shooting Date'].tolist()
                shooting_dates_set = set(shooting_dates)

                full_date_range = pd.date_range(
                    start=arrival_date + timedelta(days=1),
                    end=departure_date - timedelta(days=1)
                )
                gap_dates = [date for date in full_date_range if date not in shooting_dates_set]
                gap_days = len(gap_dates)

                # Apply weekend policy (if enabled)
                if weekend_policy:
                    # Implement any weekend-specific logic here if needed
                    pass

                total_accommodation += period_accommodation
                total_travel += period_travel

                # Collect data for export
                period_start_date = arrival_date
                period_end_date = departure_date

                # Modify description to avoid redundancy
                if home_location.lower() == (period_name or '').lower():
                    # If shooting location is same as home location, no need to mention it twice
                    description = f"{name} ({home_location}) travel tickets {period_start_date.strftime('%d.%m.%Y')}–{period_end_date.strftime('%d.%m.%Y')}"
                else:
                    description = f"{name} ({home_location}, {period_name}) travel tickets {period_start_date.strftime('%d.%m.%Y')}–{period_end_date.strftime('%d.%m.%Y')}"

                export_data.append({
                    'Cast Member': name,
                    'Home Location': home_location,
                    'Period Name': period_name if period_name else 'Local',
                    'Period Number': period_number,
                    'Period Dates': f"{period_start_date.strftime('%d.%m.%Y')}–{period_end_date.strftime('%d.%m.%Y')}",
                    'Period Start Date': period_start_date,
                    'Period End Date': departure_date,
                    'Period Travel Days': period_travel,
                    'Period Accommodation Nights': period_accommodation,
                    'Number of Shooting Days': period_length,
                    'Number of Gap Days': gap_days,
                    'Number of Travel Days': period_travel,
                    'Shooting Dates': period_group['Shooting Date'].tolist(),
                    'X': 1,  # Default value for x
                    '4X': 1,  # Default value for 4x
                    'Requires Travel': requires_travel  # Ensure the key is present
                })

                # Collect calendar events
                # Travel to location
                calendar_data.append({
                    'Cast Member': name,
                    'Title': f"Travel to {period_name}",
                    'Start': arrival_date.strftime('%d.%m.%Y'),
                    'End': arrival_date.strftime('%d.%m.%Y'),
                    'Description': f"Travel from {home_location} to {period_name}",
                    'Location': period_name,
                    'AllDay': 'TRUE'
                })

                # Shooting days
                for shooting_date in period_group['Shooting Date']:
                    calendar_data.append({
                        'Cast Member': name,
                        'Title': f"Shooting in {period_name}",
                        'Start': shooting_date.strftime('%d.%m.%Y'),
                        'End': shooting_date.strftime('%d.%m.%Y'),
                        'Description': f"Shooting day in {period_name}",
                        'Location': period_name,
                        'AllDay': 'TRUE'
                    })

                # Gap days
                for gap_date in gap_dates:
                    calendar_data.append({
                        'Cast Member': name,
                        'Title': f"Gap Day in {period_name}",
                        'Start': gap_date.strftime('%d.%m.%Y'),
                        'End': gap_date.strftime('%d.%m.%Y'),
                        'Description': f"Gap day in {period_name}",
                        'Location': period_name,
                        'AllDay': 'TRUE'
                    })

                # Travel back
                calendar_data.append({
                    'Cast Member': name,
                    'Title': f"Travel back to {home_location}",
                    'Start': departure_date.strftime('%d.%m.%Y'),
                    'End': departure_date.strftime('%d.%m.%Y'),
                    'Description': f"Travel from {period_name} to {home_location}",
                    'Location': home_location,
                    'AllDay': 'TRUE'
                })

                # Debugging
                debug_info = {
                    'Cast Member': name,
                    'Home Location': home_location,
                    'Period Name': period_name if period_name else 'Local',
                    'Period Number': period_number,
                    'Arrival Date': arrival_date.strftime('%d.%m.%Y'),
                    'Departure Date': departure_date.strftime('%d.%m.%Y'),
                    'Period Travel Days': period_travel,
                    'Period Accommodation Nights': period_accommodation,
                    'Number of Gap Days': gap_days,
                    'Total Travel Days So Far': total_travel,
                    'Total Accommodation Nights So Far': total_accommodation,
                }
                summary.append(debug_info)
            else:
                # If no travel required, assume shooting at home location
                # Shooting days
                for shooting_date in period_group['Shooting Date']:
                    calendar_data.append({
                        'Cast Member': name,
                        'Title': f"Shooting in {period_name}",
                        'Start': shooting_date.strftime('%d.%m.%Y'),
                        'End': shooting_date.strftime('%d.%m.%Y'),
                        'Description': f"Shooting day in {period_name}",
                        'Location': period_name,
                        'AllDay': 'TRUE'
                    })

                # Gap days
                # Calculate gap days within the period
                shooting_dates = period_group['Shooting Date'].tolist()
                shooting_dates_set = set(shooting_dates)

                full_date_range = pd.date_range(
                    start=period_group['Shooting Date'].min(),
                    end=period_group['Shooting Date'].max()
                )
                gap_dates = [date for date in full_date_range if date not in shooting_dates_set]

                for gap_date in gap_dates:
                    calendar_data.append({
                        'Cast Member': name,
                        'Title': f"Gap Day in {period_name}",
                        'Start': gap_date.strftime('%d.%m.%Y'),
                        'End': gap_date.strftime('%d.%m.%Y'),
                        'Description': f"Gap day in {period_name}",
                        'Location': period_name,
                        'AllDay': 'TRUE'
                    })

        # Collect results for the cast member and period
        results.append({
            'Cast Member': name,
            'Home Location': home_location,
            'Travel Days': total_travel,
            'Accommodation Nights': total_accommodation,
            'Requires Travel': requires_travel
        })

    # Create DataFrame from results
    df_summary = pd.DataFrame(results)

    # For debugging: log the summary
    for item in summary:
        logging.debug(f"Summary: {item}")

    return df_summary, export_data, calendar_data

# GUI Classes
class ManageHomeLocationsDialog(QDialog):
    def __init__(self, parent=None, home_locations=None):
        """
        Dialog to manage home locations.

        Args:
            parent (QWidget, optional): Parent widget.
            home_locations (list, optional): Existing list of home locations.
        """
        super().__init__(parent)
        self.setWindowTitle("Manage Home Locations")
        self.home_locations = home_locations if home_locations else []
        self.save_as_default = False  # New attribute to track checkbox state
        self.init_ui()

    def init_ui(self):
        """
        Initializes the UI components for managing home locations.
        """
        layout = QVBoxLayout()

        # List Widget to display home locations
        self.list_widget = QListWidget()
        self.list_widget.addItems(self.home_locations)
        layout.addWidget(self.list_widget)

        # Input field to add new location
        add_layout = QHBoxLayout()
        self.add_input = QLineEdit()
        self.add_input.setPlaceholderText("Enter new location")
        self.add_button = QPushButton("Add")
        self.add_button.clicked.connect(self.add_location)
        add_layout.addWidget(self.add_input)
        add_layout.addWidget(self.add_button)
        layout.addLayout(add_layout)

        # Remove button
        self.remove_button = QPushButton("Remove Selected")
        self.remove_button.clicked.connect(self.remove_selected)
        layout.addWidget(self.remove_button)

        # New Checkbox to save changes as default
        self.save_default_checkbox = QCheckBox("Save changes as default for new projects")
        layout.addWidget(self.save_default_checkbox)

        # Dialog buttons (OK and Cancel)
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

        self.setLayout(layout)

    def add_location(self):
        """
        Adds a new home location to the list.
        """
        location = self.add_input.text().strip()
        if location and location not in self.home_locations:
            self.list_widget.addItem(location)
            self.add_input.clear()
            logging.info(f"Added new home location: {location}")
        elif location in self.home_locations:
            QMessageBox.warning(self, "Duplicate Location", f"The location '{location}' already exists.")
            logging.warning(f"Attempted to add duplicate home location: {location}")
        else:
            QMessageBox.warning(self, "Invalid Input", "Please enter a valid location name.")
            logging.warning("Attempted to add invalid home location.")

    def remove_selected(self):
        """
        Removes the selected home location(s) from the list.
        """
        selected_items = self.list_widget.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "No Selection", "Please select a location to remove.")
            logging.warning("No home location selected for removal.")
            return
        for item in selected_items:
            location = item.text()
            self.list_widget.takeItem(self.list_widget.row(item))
            logging.info(f"Removed home location: {location}")

    def get_home_locations(self):
        """
        Retrieves the updated list of home locations.

        Returns:
            list: Updated list of home locations.
        """
        locations = []
        for index in range(self.list_widget.count()):
            item = self.list_widget.item(index)
            locations.append(item.text())
        logging.debug(f"Updated home locations: {locations}")
        return locations

    def accept(self):
        """
        Overrides the accept method to capture the checkbox state before closing.
        """
        self.save_as_default = self.save_default_checkbox.isChecked()
        super().accept()

class ManageShootingPeriodsDialog(QDialog):
    def __init__(self, parent=None, shooting_periods=None, home_locations=None):
        """
        Dialog to manage shooting periods.

        Args:
            parent (QWidget, optional): Parent widget.
            shooting_periods (list, optional): Existing list of shooting periods.
            home_locations (list, optional): List of home locations to populate the Location dropdown.
        """
        super().__init__(parent)
        self.setWindowTitle("Manage Shooting Periods")
        self.shooting_periods = shooting_periods if shooting_periods else []
        self.home_locations = home_locations if home_locations else []
        self.init_ui()

    def init_ui(self):
        """
        Initializes the UI components for managing shooting periods.
        """
        layout = QVBoxLayout()
        self.table = QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(['Name', 'Location', 'Start Date', 'End Date'])
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setSelectionMode(QTableWidget.SingleSelection)
        self.load_periods()
        layout.addWidget(self.table)

        buttons_layout = QHBoxLayout()

        add_button = QPushButton("Add Period")
        add_button.clicked.connect(self.add_period)
        buttons_layout.addWidget(add_button)

        remove_button = QPushButton("Remove Selected Period")
        remove_button.clicked.connect(self.remove_period)
        buttons_layout.addWidget(remove_button)

        layout.addLayout(buttons_layout)

        # Dialog buttons (OK and Cancel)
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

        self.setLayout(layout)

    def load_periods(self):
        """
        Loads existing shooting periods into the table.
        """
        self.table.setRowCount(len(self.shooting_periods))
        for row, period in enumerate(self.shooting_periods):
            name = period.get('Name', '')
            location = period.get('Location', '')
            start_date = period.get('Start Date')
            end_date = period.get('End Date')

            # Name
            name_item = QTableWidgetItem(name)
            self.table.setItem(row, 0, name_item)

            # Location - use QComboBox
            location_combo = QComboBox()
            location_combo.addItems(self.home_locations)
            if location in self.home_locations:
                location_combo.setCurrentText(location)
            self.table.setCellWidget(row, 1, location_combo)

            # Start Date
            start_date_edit = QDateEdit()
            start_date_edit.setDisplayFormat('dd.MM.yyyy')
            start_date_edit.setCalendarPopup(True)
            if pd.notnull(start_date):
                try:
                    start_date_parsed = pd.to_datetime(start_date)
                    if pd.isnull(start_date_parsed):
                        raise ValueError("Parsed Start Date is NaT")
                    start_date_edit.setDate(QDate(start_date_parsed.year, start_date_parsed.month, start_date_parsed.day))
                except Exception as e:
                    QMessageBox.critical(self, "Date Parsing Error", f"Error parsing Start Date for period '{name}':\n{e}")
                    start_date_edit.setDate(QDate.currentDate())
            else:
                start_date_edit.setDate(QDate.currentDate())
            self.table.setCellWidget(row, 2, start_date_edit)

            # End Date
            end_date_edit = QDateEdit()
            end_date_edit.setDisplayFormat('dd.MM.yyyy')
            end_date_edit.setCalendarPopup(True)
            if pd.notnull(end_date):
                try:
                    end_date_parsed = pd.to_datetime(end_date)
                    if pd.isnull(end_date_parsed):
                        raise ValueError("Parsed End Date is NaT")
                    end_date_edit.setDate(QDate(end_date_parsed.year, end_date_parsed.month, end_date_parsed.day))
                except Exception as e:
                    QMessageBox.critical(self, "Date Parsing Error", f"Error parsing End Date for period '{name}':\n{e}")
                    end_date_edit.setDate(QDate.currentDate())
            else:
                end_date_edit.setDate(QDate.currentDate())
            self.table.setCellWidget(row, 3, end_date_edit)

            logging.debug(f"Loaded Shooting Period - Name: {name}, Start Date: {start_date}, End Date: {end_date}")

    def add_period(self):
        """
        Adds a new shooting period to the table.
        """
        row = self.table.rowCount()
        self.table.insertRow(row)

        # Name
        name_item = QTableWidgetItem('')
        self.table.setItem(row, 0, name_item)

        # Location - use QComboBox
        location_combo = QComboBox()
        location_combo.addItems(self.home_locations)
        self.table.setCellWidget(row, 1, location_combo)

        # Start Date
        start_date_edit = QDateEdit()
        start_date_edit.setDisplayFormat('dd.MM.yyyy')
        start_date_edit.setCalendarPopup(True)
        self.table.setCellWidget(row, 2, start_date_edit)

        # End Date
        end_date_edit = QDateEdit()
        end_date_edit.setDisplayFormat('dd.MM.yyyy')
        end_date_edit.setCalendarPopup(True)
        self.table.setCellWidget(row, 3, end_date_edit)

        logging.info("Added new shooting period row.")

    def remove_period(self):
        """
        Removes the selected shooting period from the table.
        """
        selected_rows = self.table.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.warning(self, "No Selection", "Please select a period to remove.")
            logging.warning("No shooting period selected for removal.")
            return
        for selected_row in selected_rows:
            row_number = selected_row.row()
            period_name = self.table.item(row_number, 0).text()
            self.table.removeRow(row_number)
            logging.info(f"Removed shooting period: {period_name}")

    def get_shooting_periods(self):
        """
        Retrieves the updated list of shooting periods from the table.

        Returns:
            list: Updated list of shooting periods.
        """
        periods = []
        for row in range(self.table.rowCount()):
            # Name
            name_item = self.table.item(row, 0)
            name = name_item.text() if name_item else ''

            # Location
            location_widget = self.table.cellWidget(row, 1)
            location = location_widget.currentText() if isinstance(location_widget, QComboBox) else ''

            # Start Date
            start_widget = self.table.cellWidget(row, 2)
            end_widget = self.table.cellWidget(row, 3)
            if isinstance(start_widget, QDateEdit) and isinstance(end_widget, QDateEdit):
                start_qdate = start_widget.date()
                end_qdate = end_widget.date()
                try:
                    start_date = pd.to_datetime(start_qdate.toString('dd.MM.yyyy'), format='%d.%m.%Y')
                    end_date = pd.to_datetime(end_qdate.toString('dd.MM.yyyy'), format='%d.%m.%Y')
                    if pd.notnull(start_date) and pd.notnull(end_date):
                        periods.append({'Name': name, 'Location': location, 'Start Date': start_date, 'End Date': end_date})
                        logging.debug(f"Added shooting period: {name}, Location: {location}, Start: {start_date}, End: {end_date}")
                except Exception as e:
                    QMessageBox.critical(self, "Date Parsing Error", f"Error parsing dates for period '{name}':\n{e}")
                    logging.error(f"Error parsing dates for period '{name}': {e}")
                    continue
        return periods

class MainWindow(QMainWindow):
    def __init__(self):
        """
        Initializes the main window of the application.
        """
        super().__init__()
        self.setWindowTitle("Cast Travel and Accommodation Calculator")
        self.df_cast_dates = None
        self.config = {}
        self.cast_member_settings = {}
        self.shooting_periods = []
        self.calendar_data = []
        self.export_data = []
        self.current_project = None
        self.excel_file = None
        self.default_config = load_default_config()

        # Initialize menu bar and actions only once
        self.menu_bar = self.menuBar()
        self.init_menus()

        # Initialize UI
        self.init_ui()

    def init_menus(self):
        """
        Initializes the menu bar and its actions.
        """
        self.file_menu = self.menu_bar.addMenu('File')

        # New Project Action
        self.new_project_action = QAction('New Project', self)
        self.new_project_action.triggered.connect(self.new_project)
        self.file_menu.addAction(self.new_project_action)

        # Open Project Action
        self.open_project_action = QAction('Open Project', self)
        self.open_project_action.triggered.connect(self.open_project)
        self.file_menu.addAction(self.open_project_action)

        # Save Project Action
        self.save_project_action = QAction('Save Project', self)
        self.save_project_action.triggered.connect(self.save_project)
        self.file_menu.addAction(self.save_project_action)

    def init_ui(self):
        """
        Initializes the UI components of the main window.
        """
        # Main Widget and Layout
        main_widget = QWidget()
        main_layout = QVBoxLayout()
        
        # Initialize home_location_filter_list regardless of df_cast_dates
        self.home_location_filter_list = QListWidget()
        self.home_location_filter_list.setSelectionMode(QListWidget.MultiSelection)
        home_locations = self.config.get("home_locations", ["Local", "Away"])
        for location in home_locations:
            item = QListWidgetItem(location)
            item.setSelected(True)  # By default, all locations are selected
            self.home_location_filter_list.addItem(item)
        self.home_location_filter_list.itemSelectionChanged.connect(self.load_cast_members)

        # Modified condition here
        if self.df_cast_dates is None or self.df_cast_dates.empty:
            # No project loaded, display a welcome message
            welcome_label = QLabel("Welcome to the Cast Travel and Accommodation Calculator.\nPlease create a new project or open an existing one.")
            welcome_label.setAlignment(Qt.AlignCenter)
            main_layout.addWidget(welcome_label)
            main_widget.setLayout(main_layout)
            self.setCentralWidget(main_widget)
            return

        # Cast Member Table
        cast_table_layout = QVBoxLayout()
        cast_label = QLabel("Cast Members:")
        cast_table_layout.addWidget(cast_label)

        # Filtering Controls
        filter_layout = QHBoxLayout()
        self.show_included_only_checkbox = QCheckBox("Show Included Only")
        self.show_included_only_checkbox.setChecked(False)
        self.show_included_only_checkbox.stateChanged.connect(self.load_cast_members)
        filter_layout.addWidget(self.show_included_only_checkbox)

        self.home_location_filter_label = QLabel("Filter by Home Location:")
        filter_layout.addWidget(self.home_location_filter_label)

        filter_layout.addWidget(self.home_location_filter_list)

        cast_table_layout.addLayout(filter_layout)

        # Buttons for Select All / Unselect All
        select_buttons_layout = QHBoxLayout()
        self.select_all_include_button = QPushButton("Select All Include")
        self.select_all_include_button.clicked.connect(self.select_all_include)
        self.unselect_all_include_button = QPushButton("Unselect All Include")
        self.unselect_all_include_button.clicked.connect(self.unselect_all_include)
        select_buttons_layout.addWidget(self.select_all_include_button)
        select_buttons_layout.addWidget(self.unselect_all_include_button)
        cast_table_layout.addLayout(select_buttons_layout)

        # Manage Home Locations Button
        manage_locations_layout = QHBoxLayout()
        self.manage_locations_button = QPushButton("Manage Home Locations")
        self.manage_locations_button.clicked.connect(self.manage_home_locations)
        manage_locations_layout.addStretch()
        manage_locations_layout.addWidget(self.manage_locations_button)
        cast_table_layout.addLayout(manage_locations_layout)

        self.cast_table = QTableWidget()
        self.cast_table.setColumnCount(3)
        self.cast_table.setHorizontalHeaderLabels(['Include', 'Cast Member', 'Home Location'])
        self.cast_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.load_cast_members()
        cast_table_layout.addWidget(self.cast_table)

        main_layout.addLayout(cast_table_layout)

        # Summary Section
        summary_layout = QVBoxLayout()
        summary_label = QLabel("Cast Member Summary:")
        summary_layout.addWidget(summary_label)

        self.summary_text = QTextEdit()
        self.summary_text.setReadOnly(True)
        summary_layout.addWidget(self.summary_text)

        main_layout.addLayout(summary_layout)

        # Logic Parameters
        logic_layout = QHBoxLayout()

        # Max Gap Days
        max_gap_label = QLabel("Maximum Gap Days:")
        self.max_gap_spinbox = QSpinBox()
        self.max_gap_spinbox.setValue(2)
        self.max_gap_spinbox.setMinimum(0)
        self.max_gap_spinbox.setMaximum(30)
        logic_layout.addWidget(max_gap_label)
        logic_layout.addWidget(self.max_gap_spinbox)

        # Weekend Policy
        self.weekend_checkbox = QCheckBox("Send Cast Home on Weekends")
        self.weekend_checkbox.setChecked(True)
        logic_layout.addWidget(self.weekend_checkbox)

        # Arrival Option
        arrival_label = QLabel("Arrival Option:")
        self.arrival_combobox = QComboBox()
        self.arrival_combobox.addItems(["Day Before Shooting", "Same Day as Shooting"])
        logic_layout.addWidget(arrival_label)
        logic_layout.addWidget(self.arrival_combobox)

        # Departure Option
        departure_label = QLabel("Departure Option:")
        self.departure_combobox = QComboBox()
        self.departure_combobox.addItems(["Same Day as Shooting", "Day After Shooting"])
        logic_layout.addWidget(departure_label)
        logic_layout.addWidget(self.departure_combobox)

        main_layout.addLayout(logic_layout)

        # Shooting Periods Button
        self.shooting_periods_button = QPushButton("Manage Shooting Periods")
        self.shooting_periods_button.clicked.connect(self.open_shooting_periods_dialog)
        main_layout.addWidget(self.shooting_periods_button)

        # Apply Logic Button
        self.apply_button = QPushButton("Apply Logic")
        self.apply_button.clicked.connect(self.apply_logic)
        main_layout.addWidget(self.apply_button)

        # Export to Excel Button
        self.export_button = QPushButton("Export to Excel")
        self.export_button.clicked.connect(self.export_to_excel)
        main_layout.addWidget(self.export_button)

        # Export Calendar Button
        self.export_calendar_button = QPushButton("Export Calendar")
        self.export_calendar_button.clicked.connect(self.export_calendar_to_excel)
        main_layout.addWidget(self.export_calendar_button)

        # Results Table
        self.table_widget = QTableWidget()
        main_layout.addWidget(self.table_widget)

        main_widget.setLayout(main_layout)
        self.setCentralWidget(main_widget)

        # Connect signals for updating summary
        self.cast_table.itemChanged.connect(self.update_cast_member_settings)
        # Initialize settings and summary
        self.update_cast_member_settings()

    def new_project(self):
        """
        Creates a new project by prompting the user to select a cast Excel file and enter a project name.
        """
        # Prompt user to select a cast Excel file
        options = QFileDialog.Options()
        excel_file, _ = QFileDialog.getOpenFileName(self, "Select Cast Excel File", "", "Excel Files (*.xlsx);;All Files (*)", options=options)
        if not excel_file:
            logging.info("New project creation cancelled by user.")
            return

        # Prompt user to enter a project name
        project_name, ok = QInputDialog.getText(self, 'Project Name', 'Enter a name for the new project:')
        if not ok or not project_name.strip():
            QMessageBox.warning(self, "Invalid Project Name", "Project name cannot be empty.")
            return
        self.current_project = project_name.strip()

        # Load the cast data
        try:
            self.df_cast_dates = load_and_process_schedule(excel_file)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load cast Excel file:\n{e}")
            return

        self.df_cast_dates = calculate_gaps(self.df_cast_dates)

        # Initialize new configurations for the project
        self.config = {
            "home_locations": self.default_config.get("home_locations", ["Local", "Away"]),
            "shooting_periods": [],
            "cast_member_settings": {}
        }
        self.cast_member_settings = {
            name: {'include': True, 'home_location': 'Local'}
            for name in self.df_cast_dates['Cast Member'].unique()
        }
        self.config["cast_member_settings"] = self.cast_member_settings
        self.shooting_periods = []
        self.calendar_data = []
        self.export_data = []
        self.excel_file = excel_file

        # Reinitialize the UI components with the new data
        self.init_ui()
        logging.info(f"Created new project: {self.current_project}")

    def open_project(self):
        """
        Opens an existing project by prompting the user to select a project file.
        """
        # Prompt user to select a project file
        options = QFileDialog.Options()
        project_file, _ = QFileDialog.getOpenFileName(
            self,
            "Open Project File",
            "",
            "Project Files (*.json);;All Files (*)",
            options=options
        )
        if not project_file:
            logging.info("Project opening cancelled by user.")
            return

        try:
            # Load the project configuration
            project_config = load_project_config(project_file)
            self.config = project_config.get('config', {})
            self.cast_member_settings = self.config.get('cast_member_settings', {})
            self.shooting_periods = self.config.get('shooting_periods', [])
            self.current_project = project_config.get('project_name', 'Untitled')
            excel_file = project_config.get('excel_file')
            if not excel_file or not os.path.exists(excel_file):
                QMessageBox.warning(
                    self,
                    "Excel File Not Found",
                    "The cast Excel file for this project was not found."
                )
                return
            self.excel_file = excel_file
            self.df_cast_dates = load_and_process_schedule(excel_file)
            self.df_cast_dates = calculate_gaps(self.df_cast_dates)
            self.calendar_data = []
            self.export_data = []

            # Update the UI with the loaded data
            self.init_ui()
            logging.info(f"Opened project: {self.current_project}")

        except Exception as e:
            QMessageBox.critical(
                self,
                "Error",
                f"Failed to open project file:\n{e}"
            )
            logging.error(f"Failed to open project file: {e}")

        # Load the project configuration
        try:
            project_config = load_project_config(project_file)
            self.config = project_config.get('config', {})
            self.cast_member_settings = self.config.get('cast_member_settings', {})
            self.shooting_periods = self.config.get('shooting_periods', [])
            self.current_project = project_config.get('project_name', 'Untitled')
            excel_file = project_config.get('excel_file')
            if not excel_file or not os.path.exists(excel_file):
                QMessageBox.warning(self, "Excel File Not Found", "The cast Excel file for this project was not found.")
                return
            self.excel_file = excel_file
            self.df_cast_dates = load_and_process_schedule(excel_file)
            self.df_cast_dates = calculate_gaps(self.df_cast_dates)
            self.calendar_data = []
            self.export_data = []

            # Update the UI with the loaded home locations
            # Remove or comment out this block
            #self.home_location_filter_list.clear()
            #home_locations = self.config.get("home_locations", ["Local", "Away"])
            #for location in home_locations:
            #    item = QListWidgetItem(location)
            #    item.setSelected(True)
            #    self.home_location_filter_list.addItem(item)

            # Reinitialize the UI components with the new data
            self.init_ui()
            logging.info(f"Opened project: {self.current_project}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to open project file:\n{e}")
            logging.error(f"Failed to open project file: {e}")

    def save_project(self):
        """
        Saves the current project to a project file.
        """
        if not self.current_project:
            QMessageBox.warning(self, "No Project", "There is no project to save.")
            return

        # Prompt user to select a location to save the project
        options = QFileDialog.Options()
        project_file, _ = QFileDialog.getSaveFileName(self, "Save Project File", f"{self.current_project}.json", "Project Files (*.json);;All Files (*)", options=options)
        if not project_file:
            logging.info("Project saving cancelled by user.")
            return

        # Collect project data
        project_data = {
            'project_name': self.current_project,
            'excel_file': self.excel_file,
            'config': self.config
        }

        # Save to JSON file
        try:
            save_project_config(project_file, project_data)
            QMessageBox.information(self, "Project Saved", f"Project '{self.current_project}' saved successfully.")
            logging.info(f"Project '{self.current_project}' saved successfully.")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save project:\n{e}")
            logging.error(f"Failed to save project: {e}")

    def load_cast_members(self):
        """
        Loads cast members into the table with their settings, applying filters.
        """
        home_locations = self.config.get("home_locations", ["Local", "Away"])
        all_cast_members = sorted(self.df_cast_dates['Cast Member'].unique())

        # Apply filters
        # Filter by inclusion status
        if self.show_included_only_checkbox.isChecked():
            cast_members = [name for name in all_cast_members if self.cast_member_settings.get(name, {}).get('include', True)]
        else:
            cast_members = all_cast_members

        # Filter by home location
        selected_locations = [item.text() for item in self.home_location_filter_list.selectedItems()]
        cast_members = [name for name in cast_members if self.cast_member_settings.get(name, {}).get('home_location', 'Local') in selected_locations]

        self.cast_table.blockSignals(True)
        self.cast_table.setRowCount(len(cast_members))
        for row, name in enumerate(cast_members):
            settings = self.cast_member_settings.get(name, {'include': True, 'home_location': 'Local'})

            # Include Checkbox
            include_checkbox = QCheckBox()
            include_checkbox.setChecked(settings.get('include', True))
            include_checkbox.stateChanged.connect(self.update_cast_member_settings)
            self.cast_table.setCellWidget(row, 0, include_checkbox)

            # Cast Member Name
            name_item = QTableWidgetItem(name)
            name_item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
            self.cast_table.setItem(row, 1, name_item)

            # Home Location ComboBox
            location_combo = QComboBox()
            location_combo.addItems(home_locations)
            location_combo.setCurrentText(settings.get('home_location', 'Local'))
            location_combo.currentTextChanged.connect(self.update_cast_member_settings)
            self.cast_table.setCellWidget(row, 2, location_combo)

        self.cast_table.blockSignals(False)
        logging.info("Loaded cast members into the table with applied filters.")

    def update_cast_member_settings(self, item=None):
        """
        Updates the settings for each cast member based on the table inputs.
        """
        # Update settings for all cast members displayed in the table
        for row in range(self.cast_table.rowCount()):
            include_checkbox = self.cast_table.cellWidget(row, 0)
            name_item = self.cast_table.item(row, 1)
            location_combo = self.cast_table.cellWidget(row, 2)

            if include_checkbox and name_item and location_combo:
                name = name_item.text()
                include = include_checkbox.isChecked()
                home_location = location_combo.currentText()

                self.cast_member_settings[name] = {
                    'include': include,
                    'home_location': home_location
                }

        # Save updated settings to config
        self.config["cast_member_settings"] = self.cast_member_settings
        logging.debug("Updated cast member settings based on table inputs.")

        # Reload the cast members to apply any changes to filters
        self.load_cast_members()
        self.update_summary()

    def select_all_include(self):
        """
        Selects all 'Include' checkboxes.
        """
        for row in range(self.cast_table.rowCount()):
            include_checkbox = self.cast_table.cellWidget(row, 0)
            name_item = self.cast_table.item(row, 1)
            if include_checkbox and name_item:
                include_checkbox.blockSignals(True)
                include_checkbox.setChecked(True)
                include_checkbox.blockSignals(False)
                name = name_item.text()
                if name in self.cast_member_settings:
                    self.cast_member_settings[name]['include'] = True
        # Save updated settings to config
        self.config["cast_member_settings"] = self.cast_member_settings
        self.update_summary()
        logging.info("Selected all 'Include' checkboxes.")

    def unselect_all_include(self):
        """
        Unselects all 'Include' checkboxes.
        """
        for row in range(self.cast_table.rowCount()):
            include_checkbox = self.cast_table.cellWidget(row, 0)
            name_item = self.cast_table.item(row, 1)
            if include_checkbox and name_item:
                include_checkbox.blockSignals(True)
                include_checkbox.setChecked(False)
                include_checkbox.blockSignals(False)
                name = name_item.text()
                if name in self.cast_member_settings:
                    self.cast_member_settings[name]['include'] = False
        # Save updated settings to config
        self.config["cast_member_settings"] = self.cast_member_settings
        self.update_summary()
        logging.info("Unselected all 'Include' checkboxes.")

    def manage_home_locations(self):
        """
        Opens the dialog to manage home locations.
        """
        dialog = ManageHomeLocationsDialog(self, self.config.get("home_locations", ["Local", "Away"]))
        if dialog.exec_() == QDialog.Accepted:
            new_home_locations = dialog.get_home_locations()
            if "Local" not in new_home_locations:
                QMessageBox.warning(self, "Invalid Configuration", "'Local' must be included in home locations.")
                logging.warning("'Local' is missing from home locations.")
                return
            self.config["home_locations"] = new_home_locations

            # Check if user wants to save changes as default
            if dialog.save_as_default:
                self.default_config["home_locations"] = new_home_locations
                save_default_config(self.default_config)
                logging.info("Updated default home locations.")

            # Reload cast members to update Home Location dropdowns
            self.home_location_filter_list.clear()
            for location in new_home_locations:
                item = QListWidgetItem(location)
                item.setSelected(True)
                self.home_location_filter_list.addItem(item)
            self.cast_table.blockSignals(True)
            self.cast_table.clearContents()
            self.cast_table.setHorizontalHeaderLabels(['Include', 'Cast Member', 'Home Location'])
            self.load_cast_members()
            self.cast_table.blockSignals(False)
            self.update_cast_member_settings()
            logging.info("Managed home locations successfully.")

    def update_summary(self):
        """
        Updates the cast member summary based on current settings.
        """
        selected_cast = [
            name for name, settings in self.cast_member_settings.items()
            if settings['include']
        ]

        if not selected_cast:
            self.summary_text.setText("No cast member selected.")
            logging.info("No cast members selected for summary.")
            return

        summaries = []
        for name in selected_cast:
            df_member = self.df_cast_dates[self.df_cast_dates['Cast Member'] == name]
            shooting_days = len(df_member)
            shooting_dates = df_member['Shooting Date'].dt.strftime('%d.%m.%Y').tolist()
            gap_lengths = df_member['Gap'].astype(int).tolist()
            summary = f"Cast Member: {name}\n"
            summary += f"Shooting Days: {shooting_days}\n"
            summary += f"Shooting Dates: {', '.join(shooting_dates)}\n"
            if len(gap_lengths) > 1:
                summary += f"Gap Lengths: {', '.join(map(str, gap_lengths[1:]))}\n"  # Exclude first gap
            else:
                summary += f"Gap Lengths: None\n"
            summaries.append(summary)
        self.summary_text.setText('\n\n'.join(summaries))
        logging.info("Updated cast member summary.")

    def open_shooting_periods_dialog(self):
        """
        Opens the dialog to manage shooting periods.
        """
        # Remove periods with NaT values
        self.shooting_periods = [
            period for period in self.shooting_periods
            if pd.notnull(period.get('Start Date')) and pd.notnull(period.get('End Date'))
        ]
        dialog = ManageShootingPeriodsDialog(self, self.shooting_periods, self.config.get("home_locations", ["Local", "Away"]))
        if dialog.exec_() == QDialog.Accepted:
            self.shooting_periods = dialog.get_shooting_periods()
            self.config["shooting_periods"] = self.shooting_periods
            logging.info("Managed shooting periods successfully.")

    def apply_logic(self):
        """
        Applies the user-defined logic to calculate travel and accommodation needs.
        """
        if self.df_cast_dates is None or self.df_cast_dates.empty:
            QMessageBox.warning(self, "No Data", "No project data available. Please create or open a project.")
            return

        # Get user inputs
        max_gap = self.max_gap_spinbox.value()
        weekend_policy = self.weekend_checkbox.isChecked()
        arrival_option = self.arrival_combobox.currentText()
        departure_option = self.departure_combobox.currentText()

        logging.info("Applying logic with user-defined parameters.")
        logging.debug(f"Max Gap: {max_gap}, Weekend Policy: {weekend_policy}, Arrival Option: {arrival_option}, Departure Option: {departure_option}")

        try:
            # Apply logic
            self.df_summary, self.export_data, self.calendar_data = apply_user_logic(
                self.df_cast_dates, max_gap, weekend_policy, arrival_option, departure_option,
                self.cast_member_settings, self.shooting_periods
            )

            if self.export_data:
                logging.info("Logic applied successfully. Ready to export data.")
            else:
                logging.info("Logic applied successfully, but no travel data to export.")

            # Display results
            self.display_results()

            # Provide user feedback
            if self.export_data:
                QMessageBox.information(self, "Logic Applied", "Logic applied successfully. Ready to export.")
            else:
                QMessageBox.information(self, "Logic Applied", "Logic applied successfully. No travel data to export.")
        except Exception as e:
            logging.critical(f"Logic Application Error: {e}")
            QMessageBox.critical(self, "Logic Application Error", f"An error occurred while applying logic:\n{e}")

    def display_results(self):
        """
        Displays the logic application results in the results table.
        """
        if self.df_summary.empty:
            QMessageBox.information(self, "No Data", "No data available to display.")
            logging.info("No data available to display in results.")
            self.table_widget.clear()
            return

        df_display = self.df_summary.copy()

        self.table_widget.clear()
        self.table_widget.setRowCount(len(df_display))
        self.table_widget.setColumnCount(len(df_display.columns))
        self.table_widget.setHorizontalHeaderLabels(df_display.columns.tolist())

        for row_idx, row in df_display.iterrows():
            for col_idx, value in enumerate(row):
                item = QTableWidgetItem(str(value))
                self.table_widget.setItem(row_idx, col_idx, item)

        self.table_widget.resizeColumnsToContents()
        self.table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        logging.info("Displayed results in the table.")

    def export_to_excel(self):
        """
        Exports the calculated data to an Excel file, accommodating different rates per location.
        Entries are sorted chronologically based on 'Period Start Date'.
        """
        if not hasattr(self, 'export_data') or not self.export_data:
            QMessageBox.warning(self, "No Data to Export", "Please apply the logic first before exporting.")
            logging.warning("Attempted to export without any data.")
            return

        try:
            # Filter export_data to include only entries where 'Requires Travel' is True
            filtered_export_data = [item for item in self.export_data if item.get('Requires Travel', False)]

            if not filtered_export_data:
                QMessageBox.warning(self, "No Travel Data", "There is no travel data to export.")
                logging.warning("No travel data available to export after filtering.")
                return

            # Sort the export data by 'Period Start Date'
            filtered_export_data.sort(key=lambda x: x.get('Period Start Date', pd.Timestamp.min))

            # Ask the user where to save the Excel file
            options = QFileDialog.Options()
            filename, _ = QFileDialog.getSaveFileName(self, "Save Excel File", "", "Excel Files (*.xlsx);;All Files (*)", options=options)
            if not filename:
                logging.info("Export cancelled by user.")
                return

            # Proceed to create the Excel file
            wb = Workbook()
            ws = wb.active
            ws.title = "Travel and Accommodation"

            # Collect all unique locations and routes from export data
            unique_locations = set(item.get('Period Name', '') for item in filtered_export_data)
            unique_routes = set((item['Home Location'], item.get('Period Name', '')) for item in filtered_export_data)

            # Define rate labels
            rate_labels = ["Ticket Rate", "Accommodation Rate", "Per Diem Shooting Rate",
                           "Per Diem Travel Rate", "Per Diem Gap Day Rate", "Hourly Travel Rate"]

            # Write rate table header
            ws.append(["Location"] + rate_labels + ["Travel Hours per Route"])

            # Dictionary to store rate cell addresses per location
            rate_cell_refs = {}
            start_row = ws.max_row + 1

            # Write rate table for each location
            for location in unique_locations:
                ws.cell(row=start_row, column=1).value = location if location else "Local"
                rate_cell_refs[location] = {}
                for col_index, rate_label in enumerate(rate_labels, start=2):
                    rate_cell = ws.cell(row=start_row, column=col_index)
                    rate_cell.value = 0  # Default rate value is 0
                    rate_cell_refs[location][rate_label] = rate_cell.coordinate

                # Add placeholder for Travel Hours per Route
                travel_hours_cell = ws.cell(row=start_row, column=2 + len(rate_labels))
                travel_hours_cell.value = 0  # Default travel hours value is 0
                rate_cell_refs[location]['Travel Hours per Route'] = travel_hours_cell.coordinate

                start_row += 1

            # Empty row
            ws.append([])
            current_row = ws.max_row + 1

            # Define headers
            headers = ["Description", "Amt", "Unit", "x", "Currency", "Unit 2", "Rate", "Unit 3", "4x", "Unit 4", "Subtotal"]
            ws.append(headers)
            current_row += 1

            # Get max_gap value
            max_gap = self.max_gap_spinbox.value()

            # For each cast member and period, write data rows
            for item in filtered_export_data:
                cast_member = item['Cast Member']
                home_location = item['Home Location']
                period_name = item.get('Period Name', '')
                period_number = item['Period Number']
                period_dates_str = item['Period Dates']
                x = item.get('X', 1)
                four_x = item.get('4X', 1)

                location = period_name if period_name else "Local"
                # Get rate cell references for this location
                rate_cells = rate_cell_refs.get(location, {})

                # **Travel Tickets**
                description = f"{cast_member} ({home_location}, {location}) travel tickets {period_dates_str}"
                amt = 1
                unit = "return"
                rate_cell_in_formula = f"${rate_cells['Ticket Rate']}"  # Without '='
                rate_cell_value = f"={rate_cell_in_formula}"  # With '=' for cell value
                ws.append([description, amt, unit, x, "", "", rate_cell_value, "", four_x, "", ""])
                current_row = ws.max_row
                subtotal_formula = f"=B{current_row}*D{current_row}*G{current_row}*I{current_row}"
                ws.cell(row=current_row, column=11).value = subtotal_formula
                logging.debug(f"Exported Travel Tickets: {description}")

                # **Accommodation**
                description = f"{cast_member} ({home_location}, {location}) accommodation {period_dates_str}"
                amt = item['Period Accommodation Nights']
                unit = "nights"
                rate_cell_in_formula = f"${rate_cells['Accommodation Rate']}"
                rate_cell_value = f"={rate_cell_in_formula}"
                ws.append([description, amt, unit, x, "", "", rate_cell_value, "", four_x, "", ""])
                current_row = ws.max_row
                subtotal_formula = f"=B{current_row}*D{current_row}*G{current_row}*I{current_row}"
                ws.cell(row=current_row, column=11).value = subtotal_formula
                logging.debug(f"Exported Accommodation: {description}")

                # **Per Diems for Shooting Days**
                description = f"{cast_member} ({home_location}, {location}) per diems shooting days {period_dates_str}"
                amt = item['Number of Shooting Days']
                unit = "days"
                rate_cell_in_formula = f"${rate_cells['Per Diem Shooting Rate']}"
                rate_cell_value = f"={rate_cell_in_formula}"
                ws.append([description, amt, unit, x, "", "", rate_cell_value, "", four_x, "", ""])
                current_row = ws.max_row
                subtotal_formula = f"=B{current_row}*D{current_row}*G{current_row}*I{current_row}"
                ws.cell(row=current_row, column=11).value = subtotal_formula
                logging.debug(f"Exported Per Diems for Shooting Days: {description}")

                # **Per Diems for Travel Days**

                # Retrieve shooting dates from the item dictionary
                shooting_dates = item['Shooting Dates']  # List of datetime objects

                # Retrieve arrival_date from the item dictionary
                arrival_date = item['Period Start Date']
                arrival_date_str = arrival_date.strftime('%d.%m.%Y')
                description = f"{cast_member} ({home_location}, {location}) per diem travel day arrival {arrival_date_str}"

                # Check if arrival_date is a shooting day
                if arrival_date in shooting_dates:
                    amt = 0
                else:
                    amt = 1

                unit = "day"  # Since it's a single day
                rate_cell_in_formula = f"${rate_cells['Per Diem Travel Rate']}"
                rate_cell_value = f"={rate_cell_in_formula}"
                ws.append([description, amt, unit, x, "", "", rate_cell_value, "", four_x, "", ""])
                current_row = ws.max_row
                subtotal_formula = f"=B{current_row}*D{current_row}*G{current_row}*I{current_row}"
                ws.cell(row=current_row, column=11).value = subtotal_formula
                logging.debug(f"Exported Per Diem for Arrival Travel Day: {description}")

                # Retrieve departure_date from the item dictionary
                departure_date = item['Period End Date']
                departure_date_str = departure_date.strftime('%d.%m.%Y')
                description = f"{cast_member} ({home_location}, {location}) per diem travel day departure {departure_date_str}"

                # Check if departure_date is a shooting day
                if departure_date in shooting_dates:
                    amt = 0
                else:
                    amt = 1

                unit = "day"
                rate_cell_in_formula = f"${rate_cells['Per Diem Travel Rate']}"
                rate_cell_value = f"={rate_cell_in_formula}"
                ws.append([description, amt, unit, x, "", "", rate_cell_value, "", four_x, "", ""])
                current_row = ws.max_row
                subtotal_formula = f"=B{current_row}*D{current_row}*G{current_row}*I{current_row}"
                ws.cell(row=current_row, column=11).value = subtotal_formula
                logging.debug(f"Exported Per Diem for Departure Travel Day: {description}")

                # **Per Diems for Gap Days**
                if item['Number of Gap Days'] > 0 and item['Number of Gap Days'] <= max_gap:
                    description = f"{cast_member} ({home_location}, {location}) per diems gap days {period_dates_str}"
                    amt = item['Number of Gap Days']
                    unit = "days"
                    rate_cell_in_formula = f"${rate_cells['Per Diem Gap Day Rate']}"
                    rate_cell_value = f"={rate_cell_in_formula}"
                    ws.append([description, amt, unit, x, "", "", rate_cell_value, "", four_x, "", ""])
                    current_row = ws.max_row
                    subtotal_formula = f"=B{current_row}*D{current_row}*G{current_row}*I{current_row}"
                    ws.cell(row=current_row, column=11).value = subtotal_formula
                    logging.debug(f"Exported Per Diems for Gap Days: {description}")

                # **Travel Hours**
                description = f"{cast_member} ({home_location}, {location}) travel hours {period_dates_str}"
                amt = f"=${rate_cells['Travel Hours per Route']}"  # This remains the same
                unit = "hours"
                rate_cell_in_formula = f"${rate_cells['Hourly Travel Rate']}"
                rate_cell_value = f"={rate_cell_in_formula}"
                ws.append([description, amt, unit, x, "", "", rate_cell_value, "", four_x, "", ""])
                current_row = ws.max_row
                subtotal_formula = f"=B{current_row}*D{current_row}*G{current_row}*I{current_row}"
                ws.cell(row=current_row, column=11).value = subtotal_formula
                logging.debug(f"Exported Travel Hours: {description}")

            # Adjust column widths for better readability
            for col in range(1, 12):
                ws.column_dimensions[get_column_letter(col)].width = 25

            # Save the workbook
            wb.save(filename)
            QMessageBox.information(self, "Export Successful", f"Data exported to {filename}")
            logging.info(f"Exported data successfully to {filename}")
        except Exception as e:
            QMessageBox.critical(self, "Export Failed", f"An error occurred during export:\n{e}")
            logging.critical(f"Export Failed: {e}")

    def export_calendar_to_excel(self):
        """
        Exports the calendar data to an Excel file.
        """
        if not hasattr(self, 'calendar_data') or not self.calendar_data:
            QMessageBox.warning(self, "No Calendar Data", "Please apply the logic first before exporting the calendar.")
            logging.warning("Attempted to export calendar without data.")
            return

        try:
            # Ask the user where to save the Excel file
            options = QFileDialog.Options()
            filename, _ = QFileDialog.getSaveFileName(self, "Save Calendar Excel File", "", "Excel Files (*.xlsx);;All Files (*)", options=options)
            if not filename:
                logging.info("Calendar export cancelled by user.")
                return

            # Create a workbook and add sheets for each cast member
            wb = Workbook()
            cast_members = set(item['Cast Member'] for item in self.calendar_data)
            for cast_member in cast_members:
                ws = wb.create_sheet(title=cast_member[:31])  # Sheet names max length is 31
                ws.append(['Title', 'Start', 'End', 'Description', 'Location', 'AllDay'])

                # Get events for this cast member
                events = [item for item in self.calendar_data if item['Cast Member'] == cast_member]

                # Sort events by Start date
                events.sort(key=lambda x: pd.to_datetime(x['Start'], format='%d.%m.%Y'))

                for event in events:
                    ws.append([
                        event['Title'],
                        event['Start'],
                        event['End'],
                        event['Description'],
                        event['Location'],
                        event['AllDay']
                    ])

                # Adjust column widths
                for col in range(1, 7):
                    ws.column_dimensions[get_column_letter(col)].width = 20

            # Remove the default sheet if empty
            if 'Sheet' in wb.sheetnames and not wb['Sheet'].max_row > 1:
                wb.remove(wb['Sheet'])

            # Save the workbook
            wb.save(filename)
            QMessageBox.information(self, "Export Successful", f"Calendar exported to {filename}")
            logging.info(f"Calendar exported successfully to {filename}")

        except Exception as e:
            QMessageBox.critical(self, "Export Failed", f"An error occurred during calendar export:\n{e}")
            logging.critical(f"Calendar Export Failed: {e}")

# Main Application
def main():
    """
    Main function to run the application.
    """
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()